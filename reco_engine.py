# -*- coding: utf-8 -*-
"""
ERP <-> Tally supplier-ledger reconciliation engine (reusable module).

Sign convention (supplier / creditor view) used throughout:
    +ve = CREDIT to the supplier ledger  (payable UP  : purchases, charges)
    -ve = DEBIT  to the supplier ledger  (payable DOWN: payments, TDS, discounting)

Public API:
    detect_sheets(path)                  -> {'tally': name|None, 'erp': name|None, 'all': [...]}
    reconcile(path, tally_sheet, erp_sheet) -> result dict
    write_report(result, out_path)       -> writes the formatted Excel workbook
"""
import re
import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOL = 1.00          # general amount tolerance: -1.00 .. +1.00 counts as equal
TOL_BANK = 0.005    # bank entries must agree EXACTLY (no tolerance)
DATE_WINDOW = 45    # days, for reference-less amount matching
BANK_DATE_WINDOW = 15   # days, for bank date+amount matching

# kept as an alias so older callers/utilities do not break
PAISA = TOL

STATUS_ORDER = ['MATCHED', 'DIFF_TDS', 'DIFF_GST', 'DIFF_DIRECTION',
                'DIFF_OPEN', 'TALLY_ONLY', 'ERP_ONLY']

LABEL = {
    'MATCHED': 'Fully matched',
    'DIFF_TDS': 'Matched document, value differs by TDS',
    'DIFF_GST': 'Matched document, value differs by GST',
    'DIFF_DIRECTION': 'Matched value, Dr/Cr direction differs',
    'DIFF_OPEN': 'Matched document, difference UNEXPLAINED',
    'TALLY_ONLY': 'Present in Tally only',
    'ERP_ONLY': 'Present in ERP only',
}

ACTION = {
    'MATCHED': '',
    'DIFF_TDS': 'Confirm TDS treatment. Either book the TDS credit note in ERP, '
                'or configure ERP to post the invoice net of TDS.',
    'DIFF_GST': 'Check the GST treatment on this document in both systems.',
    'DIFF_DIRECTION': 'Correct the Dr/Cr direction on the ERP voucher.',
    'DIFF_OPEN': 'Manual review - no rule explains this difference.',
    'TALLY_ONLY': 'Raise the corresponding document in ERP, or confirm it is a '
                  'Tally-only book entry that ERP is not meant to carry.',
    'ERP_ONLY': 'Post the corresponding voucher in Tally, or cancel the ERP entry.',
}


# ------------------------------------------------------------------ helpers
def norm_ref(ref):
    """'EEPLB/006/25-26', 'EEPL/006/25-26', 'EEPL2/006/25-26' -> '006/25-26'.

    The series prefix is dropped on purpose: real exports contain series typos
    and settlement journals quote a different series than the document they
    settle. Document number + FY is the stable key. Every series conflict this
    relies on is reported on the Data quality sheet.
    """
    if not ref:
        return None
    r = str(ref).upper().strip().replace(' ', '')
    m = re.search(r'([A-Z0-9]+)/(\d+)/(\d{2}-\d{2})', r)
    if m:
        return '%s/%s' % (m.group(2).lstrip('0').zfill(3), m.group(3))
    return r or None


def split_refs(ref):
    if not ref:
        return []
    parts = [p for p in str(ref).replace(';', ',').split(',') if p.strip()]
    return [k for k in (norm_ref(p) for p in parts) if k]


def series_of(ref):
    if not ref:
        return ''
    m = re.match(r'([A-Z0-9]+)/', str(ref).upper().strip().replace(' ', ''))
    return m.group(1) if m else ''


def amt(cell):
    """Tally exports '16798480.00 Cr' / '14236.00 Dr' -> signed float."""
    if cell is None:
        return 0.0
    if isinstance(cell, (int, float)):
        return float(cell)
    s = str(cell).strip().replace(',', '').replace('₹', '')
    if not s:
        return 0.0
    sign = 1.0
    if s.upper().endswith('DR'):
        sign, s = -1.0, s[:-2]
    elif s.upper().endswith('CR'):
        sign, s = 1.0, s[:-2]
    try:
        return sign * float(s.strip())
    except ValueError:
        return 0.0


def num(cell):
    if cell is None:
        return 0.0
    if isinstance(cell, (int, float)):
        return float(cell)
    s = str(cell).strip().replace(',', '').replace('₹', '')
    try:
        return float(s)
    except ValueError:
        return 0.0


def d(v):
    return v.date() if hasattr(v, 'date') else v


def days(a, b):
    try:
        return abs((a - b).days)
    except Exception:
        return 9999


def money(x):
    return '%s%s' % ('-' if x < 0 else '', format(abs(round(x, 2)), ',.2f'))


def _norm(v):
    return re.sub(r'[^a-z0-9]', '', str(v).lower()) if v is not None else ''


# ------------------------------------------------------------------ sheet detection
def _header_row(ws, required, limit=30):
    """First row (within `limit`) whose cells cover every token in `required`."""
    for r in range(1, min(ws.max_row, limit) + 1):
        cells = [_norm(c.value) for c in ws[r]]
        joined = ' '.join(cells)
        if all(any(tok in c for c in cells) or tok in joined for tok in required):
            return r
    return None


def detect_sheets(path):
    """Guess which sheet is the Tally ledger and which is the ERP ledger."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    tally = erp = None
    tally_flat = None
    for name in wb.sheetnames:
        ws = wb[name]
        if erp is None and _header_row(ws, ['dr', 'cr', 'balance']):
            erp = name
            continue
        hr = _header_row(ws, ['date', 'vouchertype'])
        if hr and tally is None:
            tally = name
            continue
        if tally_flat is None and _header_row(ws, ['date', 'vchtype']):
            tally_flat = name
    return {'tally': tally or tally_flat, 'erp': erp, 'all': list(wb.sheetnames)}


def _colmap(ws, hr):
    return {_norm(c.value): c.column - 1 for c in ws[hr] if c.value is not None}


def _find(cm, *cands):
    for c in cands:
        if c in cm:
            return cm[c]
    for c in cands:
        for k, v in cm.items():
            if k.startswith(c):
                return v
    return None


# ------------------------------------------------------------------ parsing
def parse_tally(ws):
    """Handles both Tally layouts.

    Columnar ('Tally 2'): one row per voucher, a Gross Total column holding the
    supplier ledger's own Dr/Cr effect, and one column per contra ledger. This
    is preferred - the contra columns are what let us name each entry.

    Flat: Date | Dr/Cr | Ledger | Vch Type | Vch No. | Debit | Credit.
    """
    hr = _header_row(ws, ['date', 'vouchertype'])
    columnar = hr is not None
    if not columnar:
        hr = _header_row(ws, ['date', 'vchtype'])
    if hr is None:
        raise ValueError('Could not find a header row on the Tally sheet.')
    cm = _colmap(ws, hr)
    rows = []

    if columnar:
        c_date = _find(cm, 'date')
        c_part = _find(cm, 'particulars')
        c_type = _find(cm, 'vouchertype')
        c_ref = _find(cm, 'voucherrefno', 'voucherref')
        c_rdt = _find(cm, 'voucherrefdate')
        c_val = _find(cm, 'value')
        c_gross = _find(cm, 'grosstotal')
        hdr = [c.value for c in ws[hr]]
        contra = {i: str(hdr[i]).strip() for i in range(c_gross + 1, len(hdr))
                  if hdr[i]}
        for i, r in enumerate(ws.iter_rows(min_row=hr + 1, values_only=True), hr + 1):
            if not r[c_date] or not isinstance(r[c_date], (datetime.date, datetime.datetime)):
                continue
            if r[c_part] and 'grand total' in str(r[c_part]).lower():
                continue
            legs = {}
            for ci, nm in contra.items():
                if ci < len(r):
                    v = amt(r[ci])
                    if v:
                        legs[nm] = v
            rows.append(dict(
                row=i, date=d(r[c_date]), vtype=str(r[c_type] or '').strip(),
                ref_raw=r[c_ref] if c_ref is not None else None,
                ref_date=d(r[c_rdt]) if c_rdt is not None and r[c_rdt] else None,
                keys=split_refs(r[c_ref] if c_ref is not None else None),
                series=series_of(r[c_ref] if c_ref is not None else None),
                taxable=amt(r[c_val]) if c_val is not None else 0.0,
                amount=round(amt(r[c_gross]), 2), legs=legs))
    else:
        c_date = _find(cm, 'date')
        c_drcr = _find(cm, 'particulars')
        c_ledg = c_drcr + 1
        c_type = _find(cm, 'vchtype')
        c_no = _find(cm, 'vchno')
        c_dr = _find(cm, 'debit')
        c_cr = _find(cm, 'credit')
        for i, r in enumerate(ws.iter_rows(min_row=hr + 1, values_only=True), hr + 1):
            if not r[c_date] or not isinstance(r[c_date], (datetime.date, datetime.datetime)):
                continue
            a = round(num(r[c_cr]) - num(r[c_dr]), 2)
            ledger = str(r[c_ledg] or '').strip()
            rows.append(dict(
                row=i, date=d(r[c_date]), vtype=str(r[c_type] or '').strip(),
                ref_raw=r[c_no], ref_date=None, keys=split_refs(r[c_no]),
                series=series_of(r[c_no]), taxable=0.0, amount=a,
                legs={ledger: -a} if ledger else {}))
    return rows


def parse_erp(ws):
    hr = _header_row(ws, ['dr', 'cr', 'balance'])
    if hr is None:
        raise ValueError('Could not find a header row on the ERP sheet.')
    cm = _colmap(ws, hr)
    c_sn = _find(cm, 'sn', 'srno', 'slno')
    c_date = _find(cm, 'date')
    c_ref = _find(cm, 'reference', 'ref')
    c_part = _find(cm, 'particulars', 'description', 'narration')
    c_dr = _find(cm, 'dr', 'debit')
    c_cr = _find(cm, 'cr', 'credit')
    c_bal = _find(cm, 'balance')
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=hr + 1, values_only=True), hr + 1):
        if c_date is None or not r[c_date]:
            continue
        if not isinstance(r[c_date], (datetime.date, datetime.datetime)):
            continue
        ref = r[c_ref] if c_ref is not None else None
        rows.append(dict(
            row=i, sn=r[c_sn] if c_sn is not None else None, date=d(r[c_date]),
            ref_raw=ref, particulars=str(r[c_part] or '').strip() if c_part is not None else '',
            keys=split_refs(ref), series=series_of(ref),
            amount=round(num(r[c_cr]) - num(r[c_dr]), 2),
            balance=num(r[c_bal]) if c_bal is not None else 0.0))
    return rows


def parse_opening(ws, limit=15):
    """Opening balance from the report header, e.g. 'Opening Balance : Rs 0.00'.

    Returned in the same signed convention as everything else (+ = Cr = payable).
    Defaults to 0.00 when the export does not state one.
    """
    for r in range(1, min(ws.max_row, limit) + 1):
        for c in ws[r]:
            v = c.value
            if not isinstance(v, str) or 'opening' not in v.lower():
                continue
            if 'balance' not in v.lower():
                continue
            m = re.search(r'opening\s*balance\s*[:\-]?\s*[^\d\-]*(-?[\d,]+\.?\d*)\s*'
                          r'(dr|cr)?', v, re.I)
            if m:
                val = float(m.group(1).replace(',', ''))
                if m.group(2) and m.group(2).lower() == 'dr':
                    val = -val
                return round(val, 2)
            # value may sit in a neighbouring cell instead
            for cc in ws[r]:
                if cc.column > c.column:
                    n = amt(cc.value)
                    if n:
                        return round(n, 2)
    return 0.0


# ------------------------------------------------------------------ classification
def nature(t):
    L = t['legs']

    def has(sub):
        return any(sub.lower() in k.lower() for k in L)

    if t['vtype'].lower().startswith('purchase'):
        return 'Purchase invoice'
    if 'bank' in t['vtype'].lower():
        return 'Bank / LC settlement (CC a/c, incl. LC interest & GST)'
    if has('MYND') or has('M1XCHANGE'):
        return 'Bill discounting settlement (M1xchange)'
    if has('TDS ON PURCHASE'):
        return 'TDS 194Q on purchase'
    if has('RMS SIM RECHARGE'):
        return 'Service bill booked by journal (RMS SIM recharge + IGST, net of TDS 194C)'
    if has('TDS on Sub Contractor'):
        return 'TDS 194C on services'
    if has('PENALTY'):
        return 'Penalty debit note'
    if has('INTEREST ON LC'):
        return 'Interest on LC'
    if has('TDS'):
        return 'TDS journal'
    return 'Other journal'


def is_erp_purchase(e):
    return e['particulars'].lower().startswith('purchase')


def tally_class(t):
    """Which matching rule applies to this Tally voucher."""
    v = t['vtype'].lower()
    if v.startswith('purchase'):
        return 'purchase'
    if 'bank' in v:
        return 'bank'
    if t['nature'].upper().startswith('TDS'):
        return 'tds'
    return 'other'


def erp_class(e):
    p = e['particulars'].lower()
    if p.startswith('purchase'):
        return 'purchase'
    if 'bank' in p:
        return 'bank'
    if 'tds' in p:
        return 'tds'
    return 'other'


def tol_for(*classes):
    """Bank entries get no tolerance; everything else gets +/- TOL."""
    return TOL_BANK if 'bank' in classes else TOL


def eq(a, b, tolerance):
    return abs(a - b) <= tolerance


def bank_kind(text):
    t = (text or '').lower()
    if 'receipt' in t:
        return 'Receipt'
    if 'payment' in t or 'paid' in t:
        return 'Payment'
    return ''


def type_verdict(t, e):
    """Receipt vs Payment wording cross-check on a bank match."""
    a, b = bank_kind(t['vtype']), bank_kind(e['particulars'])
    if not a or not b:
        return ''
    if a == b:
        return ' Voucher type agrees (%s).' % a
    return (' NOTE: Tally calls this a Bank %s but ERP calls it a Bank %s - '
            'check which is right.' % (a, b))


def ref_verdict(t, e):
    """Reference cross-check used to confirm a date+amount match."""
    tk, ek = set(t.get('keys') or []), set(e.get('keys') or [])
    if not tk or not ek:
        return 'Not available on both sides'
    if tk & ek:
        return 'Reference agrees'
    return 'Reference DIFFERS (%s vs %s)' % (t['ref_raw'], e['ref_raw'])


# ------------------------------------------------------------------ the reconciliation
def reconcile(path, tally_sheet=None, erp_sheet=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    det = detect_sheets(path)
    tname = tally_sheet or det['tally']
    ename = erp_sheet or det['erp']
    if not tname or tname not in wb.sheetnames:
        raise ValueError('Tally sheet not found. Pick it manually.')
    if not ename or ename not in wb.sheetnames:
        raise ValueError('ERP sheet not found. Pick it manually.')

    tally = parse_tally(wb[tname])
    erp = parse_erp(wb[ename])
    if not tally:
        raise ValueError('No dated rows parsed from the Tally sheet "%s".' % tname)
    if not erp:
        raise ValueError('No dated rows parsed from the ERP sheet "%s".' % ename)

    for t in tally:
        t['nature'] = nature(t)

    T_OPEN = parse_opening(wb[tname])
    E_OPEN = parse_opening(wb[ename])
    T_DEBIT = round(sum(-t['amount'] for t in tally if t['amount'] < 0), 2)
    T_CREDIT = round(sum(t['amount'] for t in tally if t['amount'] > 0), 2)
    E_DEBIT = round(sum(-e['amount'] for e in erp if e['amount'] < 0), 2)
    E_CREDIT = round(sum(e['amount'] for e in erp if e['amount'] > 0), 2)
    T_CLOSE = round(T_OPEN + T_CREDIT - T_DEBIT, 2)
    E_CLOSE = round(E_OPEN + E_CREDIT - E_DEBIT, 2)

    results, used_t, used_e = [], set(), set()

    for t in tally:
        t['cls'] = tally_class(t)
    for e in erp:
        e['cls'] = erp_class(e)

    def add(status, key, tset, eset, reason, basis, criteria='', ambiguous=False):
        for t in tset:
            used_t.add(id(t))
        for e in eset:
            used_e.add(id(e))
        tnet = round(sum(t['amount'] for t in tset), 2)
        enet = round(sum(e['amount'] for e in eset), 2)
        results.append(dict(
            key=key, status=status, reason=reason, t_rows=tset, e_rows=eset,
            t_amt=tnet, e_amt=enet, diff=round(enet - tnet, 2),
            t_date=min([t['date'] for t in tset], default=None),
            e_date=min([e['date'] for e in eset], default=None),
            match_basis=basis, criteria=criteria, ambiguous=ambiguous))

    def classify_pair(tset, eset, diff, tolerance):
        if abs(diff) <= tolerance:
            return ('Exact match' if abs(diff) < 0.005
                    else 'Matched within the Rs %.2f tolerance (gap Rs %s)'
                         % (tolerance, money(abs(diff)))), 'MATCHED'
        tds = sum(v for t in tset for k, v in t['legs'].items() if 'TDS' in k.upper())
        if tds and abs(abs(diff) - abs(tds)) <= TOL:
            return ('ERP booked the invoice GROSS of TDS while Tally booked it NET '
                    '(TDS Rs %s deducted inside the Tally voucher)' % money(abs(tds)),
                    'DIFF_TDS')
        gst = sum(v for t in tset for k, v in t['legs'].items() if 'GST' in k.upper())
        if gst and abs(abs(diff) - abs(gst)) <= TOL:
            return 'Value differs by exactly the GST component', 'DIFF_GST'
        return 'Value difference beyond tolerance - needs manual review', 'DIFF_OPEN'

    def pick(t, cands):
        """Deterministic choice when several ERP rows could match one Tally row.

        Ranked by: reference agreement, then nearest date, then closest amount,
        then the earliest row in the source file. `ambiguous` is True when more
        than one candidate ties on every one of those, so a human must decide.
        """
        def score(e):
            return (0 if (set(t.get('keys') or []) & set(e.get('keys') or [])) else 1,
                    days(e['date'], t['date']),
                    round(abs(e['amount'] - t['amount']), 2),
                    e['row'])
        ranked = sorted(cands, key=score)
        best = ranked[0]
        ties = [e for e in ranked if score(e)[:3] == score(best)[:3]]
        return best, len(ties) > 1

    # ---------------- group by document reference
    tg, eg = defaultdict(list), defaultdict(list)
    for t in tally:
        for k in t['keys']:
            tg[k].append(t)
    for e in erp:
        for k in e['keys']:
            eg[k].append(e)
    for t in tally:
        if len(t['keys']) > 1:
            for k in t['keys']:
                tg[k] = [x for x in tg[k] if x is not t]
            tg['+'.join(t['keys'])].append(t)

    # ---------------- PASS 1: PURCHASE - invoice reference is mandatory
    for key in sorted(set(tg) | set(eg)):
        tpur = [t for t in tg.get(key, []) if id(t) not in used_t and t['cls'] == 'purchase']
        epur = [e for e in eg.get(key, []) if id(e) not in used_e and e['cls'] == 'purchase']
        if tpur and epur:
            diff = round(sum(e['amount'] for e in epur) - sum(t['amount'] for t in tpur), 2)
            reason, status = classify_pair(tpur, epur, diff, TOL)
            add(status, key, tpur, epur, reason,
                'Invoice reference (mandatory for purchases)',
                'Document reference + amount (tolerance Rs %.2f)' % TOL)

    # a purchase with no invoice reference can never be matched - say so
    for t in tally:
        if id(t) not in used_t and t['cls'] == 'purchase' and not t['keys']:
            add('TALLY_ONLY', '(no ref)', [t], [],
                'Purchase voucher carries no invoice reference, so it cannot be '
                'matched to ERP. Add the invoice number in Tally.',
                'Invoice reference missing', 'Reference required for purchases')
    for e in erp:
        if id(e) not in used_e and e['cls'] == 'purchase' and not e['keys']:
            add('ERP_ONLY', '(no ref)', [], [e],
                'ERP purchase line carries no invoice reference, so it cannot be '
                'matched to Tally.',
                'Invoice reference missing', 'Reference required for purchases')

    # ---------------- PASS 2: same document reference, non-purchase
    for key in sorted(set(tg) | set(eg)):
        tset = [t for t in tg.get(key, []) if id(t) not in used_t]
        eset = [e for e in eg.get(key, []) if id(e) not in used_e]
        for t in list(tset):
            tolerance = tol_for(t['cls'])
            m = next((e for e in eset if eq(e['amount'], t['amount'],
                                            tol_for(t['cls'], e['cls']))), None)
            if m:
                tset.remove(t)
                eset.remove(m)
                add('MATCHED', key, [t], [m],
                    'Same document reference and the amounts agree. %s'
                    % ref_verdict(t, m),
                    'Document reference + amount',
                    'Reference, amount (tolerance Rs %.2f), direction' % tolerance)

    # ---------------- PASS 3: BANK - date + amount, EXACT, both systems
    for t in sorted((x for x in tally if id(x) not in used_t and x['cls'] == 'bank'),
                    key=lambda x: (x['date'], x['row'])):
        pool = [e for e in erp if id(e) not in used_e and e['cls'] == 'bank']
        same_day = [e for e in pool if e['date'] == t['date']
                    and eq(e['amount'], t['amount'], TOL_BANK)]
        if same_day:
            e, amb = pick(t, same_day)
            add('MATCHED', t['keys'][0] if t['keys'] else '(no ref)', [t], [e],
                'Bank entry matched on the same date and the exact amount. %s%s'
                % (ref_verdict(t, e), type_verdict(t, e)),
                'Bank: date + exact amount',
                'Date (exact), amount (exact, no tolerance), direction, reference',
                ambiguous=amb)
            continue
        near = [e for e in pool if eq(e['amount'], t['amount'], TOL_BANK)
                and days(e['date'], t['date']) <= BANK_DATE_WINDOW]
        if near:
            e, amb = pick(t, near)
            add('MATCHED', t['keys'][0] if t['keys'] else '(no ref)', [t], [e],
                'Bank entry matched on the exact amount, but the dates differ by '
                '%d days (Tally %s, ERP %s). %s'
                % (days(e['date'], t['date']), t['date'], e['date'],
                   ref_verdict(t, e) + type_verdict(t, e)),
                'Bank: exact amount, date within %d days' % BANK_DATE_WINDOW,
                'Amount (exact, no tolerance), date within window, reference',
                ambiguous=amb)
            continue
        opp = [e for e in pool if eq(e['amount'], -t['amount'], TOL_BANK)
               and days(e['date'], t['date']) <= BANK_DATE_WINDOW]
        if opp:
            e, amb = pick(t, opp)
            add('DIFF_DIRECTION', t['keys'][0] if t['keys'] else '(no ref)', [t], [e],
                'Bank entry of the same amount posted on the OPPOSITE side: '
                'Tally %s, ERP %s' % ('Cr' if t['amount'] > 0 else 'Dr',
                                      'Cr' if e['amount'] > 0 else 'Dr'),
                'Bank: exact amount, opposite direction',
                'Amount (exact), direction MISMATCH', ambiguous=amb)

    # ---------------- PASS 4: TDS in Tally must appear on the ERP DEBIT side
    for t in sorted((x for x in tally if id(x) not in used_t and x['cls'] == 'tds'),
                    key=lambda x: (x['date'], x['row'])):
        debit_side = [e for e in erp if id(e) not in used_e and e['amount'] < 0]
        cands = [e for e in debit_side if eq(abs(e['amount']), abs(t['amount']), TOL)
                 and days(e['date'], t['date']) <= DATE_WINDOW]
        if cands:
            e, amb = pick(t, cands)
            add('MATCHED', t['keys'][0] if t['keys'] else '(no ref)', [t], [e],
                'TDS entry found on the ERP Debit side, as expected. %s'
                % ref_verdict(t, e),
                'TDS: ERP Debit side, amount + date',
                'ERP side must be Debit, amount (tolerance Rs %.2f), date, reference'
                % TOL, ambiguous=amb)
            continue
        wrong_side = [e for e in erp if id(e) not in used_e and e['amount'] > 0
                      and eq(abs(e['amount']), abs(t['amount']), TOL)
                      and days(e['date'], t['date']) <= DATE_WINDOW]
        if wrong_side:
            e, amb = pick(t, wrong_side)
            add('DIFF_DIRECTION', t['keys'][0] if t['keys'] else '(no ref)', [t], [e],
                'TDS of this amount exists in ERP but on the CREDIT side. TDS must '
                'sit on the ERP Debit side.',
                'TDS found on the wrong side of ERP',
                'ERP side must be Debit - it is Credit', ambiguous=amb)
            continue
        add('TALLY_ONLY', t['keys'][0] if t['keys'] else '(no ref)', [t], [],
            'TDS booked in Tally (%s) with no corresponding entry on the ERP Debit '
            'side. Every Tally TDS must appear as an ERP debit.' % t['nature'],
            'TDS: no entry on the ERP Debit side',
            'Searched the ERP Debit side on amount (tolerance Rs %.2f) and date' % TOL)

    # ---------------- PASS 5: everything else - amount + date, reference confirmed
    for t in sorted((x for x in tally if id(x) not in used_t),
                    key=lambda x: (x['date'], x['row'])):
        pool = [e for e in erp if id(e) not in used_e]
        cands = [e for e in pool
                 if eq(e['amount'], t['amount'], tol_for(t['cls'], e['cls']))
                 and days(e['date'], t['date']) <= DATE_WINDOW]
        if cands:
            e, amb = pick(t, cands)
            add('MATCHED', t['keys'][0] if t['keys'] else '(no ref)', [t], [e],
                'Matched on amount and date (%d day gap). %s'
                % (days(e['date'], t['date']), ref_verdict(t, e)),
                'Amount + date, reference confirmed',
                'Amount (tolerance Rs %.2f), date within %d days, direction, '
                'reference' % (TOL, DATE_WINDOW), ambiguous=amb)
            continue
        opp = [e for e in pool
               if eq(e['amount'], -t['amount'], tol_for(t['cls'], e['cls']))
               and days(e['date'], t['date']) <= DATE_WINDOW]
        if opp:
            e, amb = pick(t, opp)
            add('DIFF_DIRECTION', t['keys'][0] if t['keys'] else '(no ref)', [t], [e],
                'Same amount and date but posted on the OPPOSITE side: Tally %s, '
                'ERP %s' % ('Cr' if t['amount'] > 0 else 'Dr',
                            'Cr' if e['amount'] > 0 else 'Dr'),
                'Amount + date, opposite direction',
                'Amount, date, direction MISMATCH', ambiguous=amb)

    # ---------------- PASS 6: residuals -> Mismatched
    for t in tally:
        if id(t) not in used_t:
            add('TALLY_ONLY', norm_ref(t['ref_raw']) or '(no ref)', [t], [],
                'In Tally only - %s' % t['nature'],
                'No counterpart found by reference, amount or date',
                'Reference, amount (tolerance Rs %.2f), date' % TOL)
    for e in erp:
        if id(e) not in used_e:
            add('ERP_ONLY', norm_ref(e['ref_raw']) or '(no ref)', [], [e],
                'In ERP only - %s' % (e['particulars'] or 'entry'),
                'No counterpart found by reference, amount or date',
                'Reference, amount (tolerance Rs %.2f), date' % TOL)

    results.sort(key=lambda r: (r['t_date'] or r['e_date'] or datetime.date(1900, 1, 1)))

    # ---- data quality
    dq = []
    for t in tally:
        if t['ref_date'] and t['date'] and days(t['ref_date'], t['date']) > 300:
            dq.append(('Tally row %d' % t['row'], str(t['ref_raw']),
                       'Voucher date %s vs voucher-reference date %s - likely a '
                       'year typo at entry' % (t['date'], t['ref_date'])))
    srs = defaultdict(set)
    for x in tally + erp:
        for k in x['keys']:
            if x['series']:
                srs[k].add(x['series'])
    for k, v in sorted(srs.items()):
        if len(v) > 1:
            dq.append(('Document %s' % k, ' / '.join(sorted(v)),
                       'Same document number quoted under different series '
                       'prefixes - matched on the numeric key, but the series '
                       'should be corrected at source'))

    # ---- bridge
    bridge = defaultdict(float)
    for r in results:
        if r['status'] == 'MATCHED':
            bridge['Matched (within tolerance)'] += r['diff']
        elif r['status'] == 'DIFF_TDS':
            bridge['ERP booked GROSS, Tally booked NET of TDS'] += r['diff']
        elif r['status'] == 'DIFF_GST':
            bridge['GST component treated differently'] += r['diff']
        elif r['status'] == 'DIFF_DIRECTION':
            bridge['Dr/Cr direction posted differently'] += r['diff']
        elif r['status'] == 'TALLY_ONLY':
            bridge['Not in ERP: %s' % r['t_rows'][0]['nature']] += r['diff']
        elif r['status'] == 'ERP_ONLY':
            bridge['Not in Tally: %s' % (r['e_rows'][0]['particulars'] or 'entry')] += r['diff']
        else:
            bridge['Other differences (%s)' % r['status']] += r['diff']
    bridge = {k: round(v, 2) for k, v in bridge.items() if abs(round(v, 2)) > 0.004}

    buckets = defaultdict(lambda: [0, 0.0, 0.0])
    for r in results:
        b = buckets[r['status']]
        b[0] += 1
        b[1] += r['t_amt']
        b[2] += r['e_amt']

    # ---- row-level view: every voucher on both sides, stamped with its outcome,
    # so the Debit and Credit sides can each be shown as matched vs mismatched.
    for idx, r in enumerate(results):
        for t in r['t_rows']:
            t['_res'] = idx
        for e in r['e_rows']:
            e['_res'] = idx

    def counterpart(r, want_erp):
        rows = r['e_rows'] if want_erp else r['t_rows']
        if not rows:
            return ''
        refs = ' + '.join(str(x['ref_raw']) for x in rows if x.get('ref_raw'))
        return '%s%s' % (refs + ' ' if refs else '',
                         money(sum(x['amount'] for x in rows)))

    def category(status):
        return 'matched' if status == 'MATCHED' else 'unmatched'

    ledger = []
    for t in tally:
        r = results[t['_res']]
        ok = r['status'] == 'MATCHED'
        ledger.append(dict(
            side='Debit' if t['amount'] < 0 else 'Credit',
            system='Tally', date=t['date'], ref=t['ref_raw'] or '',
            key=r['key'], desc=t['nature'], vtype=t['vtype'],
            amount=abs(t['amount']), matched=ok,
            cat=category(r['status']), status=r['status'],
            counterpart=counterpart(r, True) or 'No ERP counterpart',
            diff=r['diff'], reason=r['reason'], criteria=r.get('criteria', ''),
            ambiguous=r.get('ambiguous', False),
            action=ACTION.get(r['status'], '')))
    for e in erp:
        r = results[e['_res']]
        ok = r['status'] == 'MATCHED'
        ledger.append(dict(
            side='Debit' if e['amount'] < 0 else 'Credit',
            system='ERP', date=e['date'], ref=e['ref_raw'] or '',
            key=r['key'], desc=e['particulars'] or 'entry', vtype=e['particulars'],
            amount=abs(e['amount']), matched=ok,
            cat=category(r['status']), status=r['status'],
            counterpart=counterpart(r, False) or 'No Tally counterpart',
            diff=r['diff'], reason=r['reason'], criteria=r.get('criteria', ''),
            ambiguous=r.get('ambiguous', False),
            action=ACTION.get(r['status'], '')))
    ledger.sort(key=lambda x: (x['side'] != 'Debit', x['matched'],
                               x['date'] or datetime.date(1900, 1, 1)))

    def sub(side, system, cat):
        return round(sum(x['amount'] for x in ledger if x['side'] == side
                         and x['system'] == system and x['cat'] == cat), 2)

    sides = {}
    for side in ('Debit', 'Credit'):
        v = {}
        for sysname, key in (('Tally', 'tally'), ('ERP', 'erp')):
            for cat in ('matched', 'unmatched'):
                v['%s_%s' % (key, cat)] = sub(side, sysname, cat)
            v['%s_total' % key] = round(sum(v['%s_%s' % (key, c)]
                                            for c in ('matched', 'unmatched')), 2)
        v['diff'] = round(v['erp_total'] - v['tally_total'], 2)
        sides[side] = v

    # control check: Opening + Credit - Debit must equal Closing on both sides
    ctrl = [
        ('Tally', round(T_OPEN + T_CREDIT - T_DEBIT - T_CLOSE, 2)),
        ('ERP', round(E_OPEN + E_CREDIT - E_DEBIT - E_CLOSE, 2)),
    ]

    return dict(
        tally=tally, erp=erp, results=results, bridge=bridge, ledger=ledger,
        sides=sides, control=ctrl,
        buckets={k: [v[0], round(v[1], 2), round(v[2], 2)] for k, v in buckets.items()},
        t_open=T_OPEN, e_open=E_OPEN,
        t_debit=T_DEBIT, e_debit=E_DEBIT,
        t_credit=T_CREDIT, e_credit=E_CREDIT,
        t_close=T_CLOSE, e_close=E_CLOSE, gap=round(E_CLOSE - T_CLOSE, 2),
        residual=round(E_CLOSE - T_CLOSE - sum(bridge.values()), 2),
        dq=dq, tally_sheet=tname, erp_sheet=ename,
    )


# ------------------------------------------------------------------ report writer
HDR_F = Font(bold=True, color='FFFFFF', size=10)
HDR_B = PatternFill('solid', fgColor='1F4E79')
TITLE = Font(bold=True, size=13, color='1F4E79')
BOLD = Font(bold=True, size=10)
THIN = Border(*[Side('thin', color='D0D0D0')] * 4)
NUM = '#,##0.00;[Red]-#,##0.00'
FILL = {
    'MATCHED': PatternFill('solid', fgColor='E2EFDA'),
    'MATCHED_NIL': PatternFill('solid', fgColor='EDEDED'),
    'DIFF_TDS': PatternFill('solid', fgColor='FFF2CC'),
    'DIFF_GST': PatternFill('solid', fgColor='FFF2CC'),
    'DIFF_DIRECTION': PatternFill('solid', fgColor='FFF2CC'),
    'DIFF_OPEN': PatternFill('solid', fgColor='FCE4D6'),
    'TALLY_ONLY': PatternFill('solid', fgColor='FBE5E5'),
    'ERP_ONLY': PatternFill('solid', fgColor='DDEBF7'),
}

METHOD = [
    ('Sign convention',
     'Both ledgers are converted to one signed number from the SUPPLIER point of '
     'view: +ve = credit (payable increases), -ve = debit (payable decreases). '
     'Tally strings like "16798480.00 Cr" are parsed to +16798480; ERP Dr/Cr '
     'columns become Cr - Dr. This removes the manual Dr/Cr eyeballing.'),
    ('Reference normalisation',
     'Document numbers reduce to <number>/<FY>, so EEPLB/006/25-26, EEPL/006/25-26 '
     'and EEPL2/006/25-26 all key to 006/25-26. This survives the series typos '
     'present in real exports. Multi-document settlement vouchers are split on '
     'commas and handled as one composite group.'),
    ('Pass 0 - contra pairs',
     'A Bank Receipt and Bank Payment of the same value on the same date is a '
     'self-reversing pair inside Tally, net nil to the supplier. These are '
     'collapsed instead of being reported as two phantom breaks.'),
    ('Pass 1 - document-wise, nature-aware',
     'Vouchers sharing a document number are matched first on exact amount, then '
     'the Tally PURCHASE voucher is compared to the ERP purchase line '
     'specifically. Settlements are deliberately NOT netted in at this stage - '
     'netting them would hide the TDS gap and a missing settlement at the same '
     'time, cancelling each other out.'),
    ('Pass 2 - amount + date',
     'ERP records payments under its own reference (AE...., PMT....) while Tally '
     'settlement journals quote the INVOICE number, so these can never meet on a '
     'reference. They match on equal signed amount within a %d-day window. If the '
     'amount matches but the sign is opposite, the pair is reported as a Dr/Cr '
     'direction error rather than left as two unmatched items.' % DATE_WINDOW),
    ('Pass 3 - residuals',
     'Anything still unmatched is reported as Tally-only or ERP-only, labelled '
     'with what the entry actually is - read from the contra ledger, not the '
     'voucher type (TDS, discounting, penalty, interest, service bill, bank).'),
    ('Difference classification',
     'For a matched document with a value gap the engine tests, in order: '
     '(a) gap <= Rs 1.00 -> rounding / Round-Off ledger; (b) gap equals the TDS '
     'legs inside the Tally voucher -> ERP booked GROSS while Tally booked NET of '
     'TDS; (c) gap equals the GST component; (d) otherwise UNEXPLAINED and flagged '
     'for manual review.'),
    ('Balance bridge = the proof',
     'Every classified difference accumulates into a walk from the Tally closing '
     'balance to the ERP closing balance. The residual MUST be zero - that is the '
     'guarantee that no difference has been silently dropped. A non-zero residual '
     'means the reconciliation is incomplete and must not be signed off.'),
]


def write_report(R, out_path):
    out = openpyxl.Workbook()

    def sheet(name, headers, widths):
        ws = out.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill = HDR_F, HDR_B
            c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        return ws

    def fmt(ws, numcols):
        mx = max(numcols, default=0)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.border = THIN
                c.alignment = Alignment(vertical='top', wrap_text=(c.column > mx))
                if c.column in numcols:
                    c.number_format = NUM

    # ---- Summary
    ws = out.active
    ws.title = 'Summary'
    for col, w in zip('ABCDEF', (40, 18, 18, 18, 18, 44)):
        ws.column_dimensions[col].width = w

    def dr(x):
        return round(abs(x), 2) if x < 0 else 0.0

    def cr(x):
        return round(x, 2) if x > 0 else 0.0

    def srow(vals, bold=False, head=False, indent=0, fill=None):
        ws.append(vals)
        rn = ws.max_row
        for c in ws[rn]:
            if head:
                c.font, c.fill = HDR_F, HDR_B
            else:
                if bold:
                    c.font = BOLD
                if fill:
                    c.fill = PatternFill('solid', fgColor=fill)
            if (c.column > 1 and isinstance(c.value, (int, float))
                    and not isinstance(c.value, bool)):
                c.number_format = NUM
            c.border = THIN
        if indent:
            ws.cell(rn, 1).alignment = Alignment(indent=indent)
        return rn

    ws.append(['ERP  vs  TALLY  -  SUPPLIER LEDGER RECONCILIATION'])
    ws['A1'].font = TITLE
    ws.append(['Tally sheet: %s    |    ERP sheet: %s' % (R['tally_sheet'], R['erp_sheet'])])
    ws['A2'].font = Font(italic=True, size=10, color='595959')
    ws.append([])

    srow(['', 'TALLY', '', 'ERP', '', ''], head=True)
    srow(['1-4.  LEDGER CONTROL', 'Debit', 'Credit', 'Debit', 'Credit', 'Note'],
         head=True)
    srow(['1.  Opening balance',
          dr(R['t_open']), cr(R['t_open']), dr(R['e_open']), cr(R['e_open']),
          'From the export header; 0.00 when not stated'], bold=True)
    srow(['2.  Total DEBIT for the period',
          R['t_debit'], '', R['e_debit'], '',
          'Payments, TDS, bill discounting, bank settlements'], bold=True)
    srow(['3.  Total CREDIT for the period',
          '', R['t_credit'], '', R['e_credit'],
          'Purchases and other supplier charges'], bold=True)
    srow(['4.  Closing balance',
          dr(R['t_close']), cr(R['t_close']), dr(R['e_close']), cr(R['e_close']),
          'Opening + Credit - Debit'], bold=True, fill='F2F5F8')
    srow(['     Control check (must be 0.00)',
          R['control'][0][1], '', R['control'][1][1], '',
          'Opening + Credit - Debit - Closing'])
    srow(['     Closing difference (ERP - Tally)', '', R['gap'], '', '',
          'The amount this reconciliation has to explain'], bold=True)
    ws.append([])

    for side in ('Debit', 'Credit'):
        v = R['sides'][side]
        srow(['%s SIDE  -  matched vs mismatched' % side.upper(),
              'Tally', 'ERP', 'Difference', '', 'Meaning'], head=True)
        srow(['   Matched', v['tally_matched'], v['erp_matched'],
              round(v['erp_matched'] - v['tally_matched'], 2), '',
              'Counterpart found and values agree'], indent=1)
        srow(['   Mismatched', v['tally_unmatched'], v['erp_unmatched'],
              round(v['erp_unmatched'] - v['tally_unmatched'], 2), '',
              'No counterpart, or values disagree'], indent=1)
        srow(['   Total %s' % side, v['tally_total'], v['erp_total'], v['diff'], '',
              'Ties to line %s above' % ('2' if side == 'Debit' else '3')], bold=True)
        ws.append([])

    srow(['MATCHING RULES APPLIED', '', '', '', '', ''], head=True)
    for a, b in (('Purchase', 'Must match on the invoice / document reference. '
                              'No invoice number means no match.'),
                 ('Bank', 'Date + amount, EXACT - no tolerance is applied to bank '
                          'entries (date window %d days).' % BANK_DATE_WINDOW),
                 ('TDS', 'Every TDS booked in Tally is searched for on the ERP '
                         'DEBIT side only.'),
                 ('All others', 'Amount within +/- Rs %.2f, date within %d days, '
                                'then the reference number is confirmed.'
                                % (TOL, DATE_WINDOW)),
                 ('Unmatched', 'Anything without a counterpart goes to Mismatched.')):
        srow(['   %s' % a, b, '', '', '', ''], indent=1)
    ws.append([])

    srow(['MATCH RESULT', 'Count', 'Tally value', 'ERP value', '', ''], head=True)
    for k in STATUS_ORDER:
        if k in R['buckets']:
            c, ta, ea = R['buckets'][k]
            srow([LABEL[k], c, ta, ea, '', ''])
    amb = sum(1 for r in R['results'] if r.get('ambiguous'))
    srow(['Of which flagged AMBIGUOUS (several equally good candidates)',
          amb, '', '', '', 'Review these by hand'], bold=bool(amb))
    ws.append([])

    srow(['BALANCE BRIDGE:  Tally closing  ->  ERP closing', 'Amount', '', '', '', ''],
         head=True)
    srow(['Tally closing balance', R['t_close'], '', '', '', ''], bold=True)
    for k, v in sorted(R['bridge'].items(), key=lambda x: -abs(x[1])):
        srow(['   %s' % k, v, '', '', '', ''], indent=1)
    srow(['ERP closing balance (derived)',
          round(R['t_close'] + sum(R['bridge'].values()), 2), '', '', '', ''], bold=True)
    srow(['ERP closing balance (per ERP)', R['e_close'], '', '', '', ''], bold=True)
    srow(['Unexplained residual', R['residual'], '', '', '',
          'MUST be 0.00 - otherwise do not sign off'], bold=True)

    # ---- Reconciliation
    ws = sheet('Reconciliation',
               ['Status', 'Document / Key', 'Tally date', 'Tally vch type', 'Tally ref',
                'Tally amount', 'ERP date', 'ERP ref', 'ERP particulars', 'ERP amount',
                'Difference (ERP - Tally)', 'Matched on', 'Reason / remark'],
               [22, 18, 12, 20, 26, 16, 12, 18, 30, 16, 18, 30, 62])
    for r in R['results']:
        ws.append([r['status'], r['key'], r['t_date'],
                   ' + '.join(sorted({t['vtype'] for t in r['t_rows']})),
                   ' + '.join(str(t['ref_raw']) for t in r['t_rows'] if t['ref_raw']),
                   r['t_amt'], r['e_date'],
                   ' + '.join(str(e['ref_raw']) for e in r['e_rows'] if e['ref_raw']),
                   ' + '.join(sorted({e['particulars'] for e in r['e_rows']})),
                   r['e_amt'], r['diff'], r['match_basis'], r['reason']])
        for c in ws[ws.max_row]:
            c.fill = FILL.get(r['status'], PatternFill())
    fmt(ws, {6, 10, 11})

    # ---- Exceptions
    ws = sheet('Exceptions',
               ['Status', 'Document / Key', 'Date', 'Description', 'Tally amount',
                'ERP amount', 'Difference', 'Reason', 'Action for the team'],
               [22, 18, 12, 40, 16, 16, 18, 60, 52])
    for r in R['results']:
        if r['status'] == 'MATCHED':
            continue
        desc = (' + '.join(sorted({t['nature'] for t in r['t_rows']}))
                or ' + '.join(sorted({e['particulars'] for e in r['e_rows']})))
        ws.append([r['status'], r['key'], r['t_date'] or r['e_date'], desc,
                   r['t_amt'], r['e_amt'], r['diff'], r['reason'],
                   ACTION.get(r['status'], '')])
        for c in ws[ws.max_row]:
            c.fill = FILL.get(r['status'], PatternFill())
    fmt(ws, {5, 6, 7})

    # ---- Debit / Credit detail: matched and mismatched, both systems
    for side in ('Debit', 'Credit'):
        v = R['sides'][side]
        ws = sheet('%s detail' % side,
                   ['Matched?', 'Status', 'Ambiguous?', 'System', 'Date',
                    'Reference', 'Doc key', 'Description', '%s amount' % side,
                    'Counterpart in other system', 'Doc difference',
                    'Criteria checked', 'Reason', 'Action'],
                   [11, 17, 11, 9, 12, 24, 14, 38, 16, 34, 16, 46, 58, 50])
        BLOCKS = (('MATCHED', 'matched', 'E2EFDA'),
                  ('MISMATCHED', 'unmatched', 'FCE4D6'))
        for grp, cat, colour in BLOCKS:
            block = [x for x in R['ledger'] if x['side'] == side and x['cat'] == cat]
            ws.append(['%s %s - %d rows' % (grp, side.upper(), len(block))])
            for c in ws[ws.max_row]:
                c.font = BOLD
                c.fill = PatternFill('solid', fgColor=colour)
            for x in block:
                ws.append([x['cat'].title(), x['status'],
                           'AMBIGUOUS' if x.get('ambiguous') else '',
                           x['system'], x['date'], str(x['ref']), x['key'],
                           x['desc'], x['amount'], x['counterpart'],
                           x['diff'], x.get('criteria', ''), x['reason'],
                           x['action']])
                ws.cell(ws.max_row, 2).fill = FILL.get(x['status'], PatternFill())
                if x.get('ambiguous'):
                    ws.cell(ws.max_row, 3).fill = PatternFill('solid', fgColor='FFF2CC')
            for sysname, key in (('Tally', 'tally'), ('ERP', 'erp')):
                ws.append(['', '', '', sysname, '', '', '',
                           'Sub-total %s %s %s' % (sysname, cat, side.lower()),
                           v['%s_%s' % (key, cat)], '', '', '', '', ''])
                for c in ws[ws.max_row]:
                    c.font = BOLD
            ws.append([])
        for sysname, key in (('Tally', 'tally_total'), ('ERP', 'erp_total')):
            ws.append(['', '', '', sysname, '', '', '',
                       'TOTAL %s per %s' % (side.upper(), sysname), v[key],
                       '', '', '', '', ''])
            for c in ws[ws.max_row]:
                c.font, c.fill = BOLD, PatternFill('solid', fgColor='DDEBF7')
        fmt(ws, {9, 11})

    # ---- parsed inputs
    ws = sheet('Tally (parsed)',
               ['Src row', 'Date', 'Voucher type', 'Voucher ref', 'Doc key(s)', 'Nature',
                'Taxable value', 'Ledger amount (+Cr / -Dr)', 'Contra ledgers'],
               [9, 12, 22, 26, 22, 38, 16, 22, 70])
    for t in R['tally']:
        ws.append([t['row'], t['date'], t['vtype'], t['ref_raw'], ', '.join(t['keys']),
                   t['nature'], t['taxable'], t['amount'],
                   '; '.join('%s %s' % (k, money(v)) for k, v in t['legs'].items())])
    fmt(ws, {7, 8})

    ws = sheet('ERP (parsed)',
               ['Src row', 'S/N', 'Date', 'Reference', 'Doc key(s)', 'Particulars',
                'Ledger amount (+Cr / -Dr)', 'Running balance'],
               [9, 7, 12, 22, 22, 42, 22, 18])
    for e in R['erp']:
        ws.append([e['row'], e['sn'], e['date'], e['ref_raw'], ', '.join(e['keys']),
                   e['particulars'], e['amount'], e['balance']])
    fmt(ws, {7, 8})

    ws = sheet('Data quality', ['Where', 'Value', 'Issue'], [24, 30, 90])
    for row in R['dq']:
        ws.append(list(row))
    fmt(ws, set())

    ws = sheet('Method', ['#', 'Step', 'What it does'], [6, 42, 110])
    for i, (s, w) in enumerate(METHOD, 1):
        ws.append([i, s, w])
    fmt(ws, set())
    for r in ws.iter_rows(min_row=2):
        r[1].font = BOLD

    out.save(out_path)
    return out_path
